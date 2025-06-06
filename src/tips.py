import calendar
import datetime
import logging
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from io import BytesIO
from itertools import islice
from locale import LC_NUMERIC, setlocale
from operator import itemgetter
from typing import Any, cast

import boto3
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.worksheet.worksheet import Worksheet
from wheniwork import WhenIWork

import flexepos
from ssm_parameter_store import SSMParameterStore

logger = logging.getLogger(__name__)

WHENIWORK_DATE_FORMAT = "%a, %d %b %Y %H:%M:%S %z"

# warning! this won't work if we multiply
TWO_PLACES = Decimal(10) ** -2
setlocale(LC_NUMERIC, "")


class Tips:
    """ """

    def __init__(self):
        self._parameters: SSMParameterStore = cast(
            SSMParameterStore, SSMParameterStore(prefix="/prod")["wheniwork"]
        )
        self._a = WhenIWork()
        self._a.login(
            self._parameters["user"],
            self._parameters["password"],
            self._parameters["key"],
        )
        self._locations: dict[int, Any] = {}
        self._stores: dict[str, Any] = {}
        locations_dict: dict[str, Any] | None = self._a.get("/locations")
        if locations_dict:
            for location in locations_dict["locations"]:
                self._locations[location["id"]] = location
                self._stores[location["name"]] = location
        return

    def payperiod_dates(self, pay_period, year_month_date):
        if pay_period == 0:
            span_dates = [
                datetime.date(year_month_date.year, year_month_date.month, 1),
                datetime.date(
                    year_month_date.year,
                    year_month_date.month,
                    calendar.monthrange(year_month_date.year, year_month_date.month)[1],
                )
                + datetime.timedelta(days=1),
            ]
        elif pay_period == 1:
            span_dates = [
                datetime.date(year_month_date.year, year_month_date.month, 1),
                datetime.date(year_month_date.year, year_month_date.month, 16),
            ]  # not inclusive on end date
        elif pay_period == 2:
            span_dates = [
                datetime.date(year_month_date.year, year_month_date.month, 16),
                datetime.date(
                    year_month_date.year,
                    year_month_date.month,
                    calendar.monthrange(year_month_date.year, year_month_date.month)[1],
                )
                + datetime.timedelta(days=1),
            ]
        else:
            raise ValueError(
                f"pay_period must be set to 0 for entire month or 1 or 2 not {pay_period}"
            )
        return span_dates

    def getTimes(self, stores, year_month_date, pay_period=0):
        span_dates = self.payperiod_dates(pay_period, year_month_date)
        self._times: dict[str, Any] | None = self._a.get(
            "/times",
            params={
                "start": span_dates[0].isoformat(),
                "end": span_dates[1].isoformat(),
                "location_id": ",".join(
                    [
                        str(self._stores[key]["id"])
                        for key in stores
                        if key in self._stores
                    ]
                ),
            },
        )

        rv = {}
        if self._times:
            for store in stores:
                rv[store] = {}
                for time in list(
                    filter(
                        lambda rtime: rtime["location_id"] == self._stores[store]["id"],
                        self._times["times"],
                    )
                ):
                    user_times = rv[store].get(time["user_id"], [])
                    user_times.append(time)
                    rv[store][time["user_id"]] = user_times

        return rv

    def attendanceReport(self, stores, start_date, end_date):
        self._shifts: dict[str, Any] | None = self._a.get(
            "/shifts",
            params={
                "start": start_date.isoformat(),
                "end": end_date.isoformat(),
                "location_id": ",".join(
                    [
                        str(self._stores[key]["id"])
                        for key in stores
                        if key in self._stores
                    ]
                ),
            },
        )

        self._times: dict[str, Any] | None = self._a.get(
            "/times",
            params={
                "start": start_date.isoformat(),
                "end": end_date.isoformat(),
                "location_id": ",".join(
                    [
                        str(self._stores[key]["id"])
                        for key in stores
                        if key in self._stores
                    ]
                ),
            },
        )

        text_csv = []

        text_csv.append(
            "Store,Name,Shift Time,Clock-in Time,Late/Early,Type".split(",")
        )

        if self._times and self._shifts:
            shift_lookup = {entry["id"]: entry for entry in self._shifts["shifts"]}
            users_lookup = {entry["id"]: entry for entry in self._shifts["users"]}
            done = defaultdict(list)

            for time in self._times["times"]:
                if time["shift_id"] in shift_lookup:
                    shift_start = datetime.datetime.strptime(
                        shift_lookup[time["shift_id"]]["start_time"],
                        WHENIWORK_DATE_FORMAT,
                    )
                    clock_in = datetime.datetime.strptime(
                        time["start_time"], WHENIWORK_DATE_FORMAT
                    )
                    user = users_lookup[time["user_id"]]
                    hash_str = str(time["user_id"]) + str(shift_start.date())
                    # find the earliest clock in for this person/date
                    if hash_str in done:
                        clock_in = min([clock_in] + done[hash_str])
                    if abs((shift_start - clock_in).total_seconds() // 60) > 5:
                        minutes = (shift_start - clock_in).total_seconds() // 60
                        text_csv.append(
                            [
                                self._locations[time["location_id"]]["name"],
                                f"{user['last_name']}, {user['first_name']}",
                                f"{shift_start.strftime('%Y-%m-%d %H:%M:%S')}",
                                f"{clock_in.strftime('%Y-%m-%d %H:%M:%S')}",
                                str(minutes),
                                "early" if minutes > 0 else "late",
                            ]
                        )
                    done[hash_str].append(shift_start)
                else:
                    # print(f"no shift for time {time['shift_id']}")
                    pass

            missed_shifts = set(shift_lookup.keys()).difference(
                {x["shift_id"] for x in self._times["times"]}
            )
            for shift_id in missed_shifts:
                shift = shift_lookup[shift_id]
                user = users_lookup[shift["user_id"]]
                text_csv.append(
                    [
                        self._locations[shift["location_id"]]["name"],
                        f"{user['last_name']}, {user['first_name']}",
                        f'"{datetime.datetime.strptime(shift["start_time"], WHENIWORK_DATE_FORMAT).strftime("%Y-%m-%d %H:%M:%S")}"',
                        "N/A",
                        "N/A",
                        "no show on shift",
                    ]
                )
        return text_csv

    def emailTips(self, stores, tip_date, pay_period=0):
        span_dates = self.payperiod_dates(pay_period, tip_date)
        parameters = cast(SSMParameterStore, SSMParameterStore(prefix="/prod")["email"])
        receiver_email = ["info@wagonermanagement.com"]
        from_email = cast(str, parameters["from_email"])
        subject = "Tip Spreadsheet for {} pp {}".format(
            tip_date.strftime("%m/%Y"), pay_period
        )
        # charset = "UTF-8"
        output = BytesIO()
        workbook = openpyxl.Workbook()
        active = workbook.active
        workbook.remove(active) if active else None
        f = flexepos.Flexepos()
        # remove the non inclusive dates as flexepos is inclusive on the end
        tip_totals = f.getTips(
            stores, span_dates[0], span_dates[1] - datetime.timedelta(days=1)
        )
        times = self.getTimes(
            stores, datetime.date(tip_date.year, tip_date.month, 1), pay_period
        )

        for store in stores:
            sheet = workbook.create_sheet(title=store)
            i = 3
            sheet.append(
                ["{} - {} pp {}".format(store, tip_date.strftime("%B"), pay_period)]
                + tip_totals[store][0]
            )
            sheet.append(["=SUM(B2:K2)"] + tip_totals[store][1])
            for n in range(1, 13):
                sheet.cell(2, n).number_format = '"$"#,##0.00_-'
            sheet.append(["Last Name", "First Name", "Hours", "Tip Share"])
            for user_times in times[store].values():
                user = self.getUser(user_times[0]["user_id"])
                if user is None:
                    continue
                first_name = user["first_name"]
                last_name = user["last_name"]
                sheet.append(
                    [
                        last_name,
                        first_name,
                        sum(item["length"] for item in user_times),
                        "=$A$2 / SUM($C$4:$C$99) * $C{}".format(i + 1),
                    ]
                )
                sheet.cell(i + 1, 4).number_format = '"$"#,##0.00_-'
                i = i + 1

            sheet.auto_filter.ref = "A3:D{}".format(i)
            sheet.auto_filter.add_sort_condition("A3:A{}".format(i))
            dim_holder = DimensionHolder(worksheet=sheet)
            for col in range(sheet.min_column, sheet.max_column + 1):
                dim_holder[get_column_letter(col)] = ColumnDimension(
                    sheet, min=col, max=col, bestFit=True
                )
            sheet.column_dimensions = dim_holder

        workbook.save(output)

        msg = MIMEMultipart()
        msg["Subject"] = subject
        msg["From"] = from_email
        msg["To"] = ", ".join(receiver_email)
        part = MIMEText("Attached is the spreadsheet\n\n")
        msg.attach(part)
        part = MIMEApplication(output.getvalue())
        part.add_header("Content-Disposition", "attachment", filename="tips.xlsx")
        msg.attach(part)

        client = boto3.client("ses")
        client.send_raw_email(
            Source=msg["From"],
            Destinations=receiver_email,  ## passed in an array
            RawMessage={
                "Data": msg.as_string(),
            },
        )

    def getMissingPunches(self) -> list[dict[str, Any]]:
        times: dict[str, Any] | None = self._a.get(
            "/times",
            params={
                "start": (
                    datetime.date.today() - datetime.timedelta(days=20)
                ).isoformat(),
                "end": (datetime.datetime.now()).isoformat(),
            },
        )
        self._users = {}
        if times:
            for user in times["users"]:
                self._users[user["id"]] = user
            return list(filter(lambda rtime: rtime["end_time"] is None, times["times"]))
        else:
            return []

    def getMealPeriodViolations(self, stores, year_month_date, pay_period=0):
        times = self.getTimes(stores, year_month_date, pay_period)
        mpvs = []
        for store in stores:
            for user_times in times[store].values():
                user = self.getUser(user_times[0]["user_id"])
                if user is None:
                    continue
                first_name = user["first_name"]
                last_name = user["last_name"]
                hourly_rate = user["hourly_rate"]
                day_dict = defaultdict(list)
                for time in user_times:
                    day = datetime.datetime.strptime(
                        time["start_time"], WHENIWORK_DATE_FORMAT
                    ).date()
                    day_dict[day].append(time)
                for day, day_times in day_dict.items():
                    day_hours = sum(i["length"] for i in day_times)
                    day_times[0]["first_name"] = first_name
                    day_times[0]["last_name"] = last_name

                    day_times[0]["store"] = store
                    day_times[0]["hourly_rate"] = hourly_rate
                    day_times[0]["day"] = day
                    if len(day_times) == 1:
                        if day_hours > 6:
                            mpvs.append(day_times[0])
                    elif day_hours > 6:
                        if day_times[0]["length"] > 5:
                            mpvs.append(day_times[0])
                        # does not take into account 12 hour shifts
                        end_time = datetime.datetime.strptime(
                            day_times[0]["end_time"], WHENIWORK_DATE_FORMAT
                        )
                        next_time = datetime.datetime.strptime(
                            day_times[1]["start_time"], WHENIWORK_DATE_FORMAT
                        )
                        break_duration = next_time - end_time
                        if break_duration < datetime.timedelta(minutes=30):
                            logger.warning(
                                "Lunch break less than 30 minutes",
                                extra=dict(
                                    day=day.isoformat(),
                                    last_name=last_name,
                                    first_name=first_name,
                                    break_duration=break_duration.total_seconds() / 60,
                                ),
                            )

        return sorted(mpvs, key=itemgetter("last_name", "first_name"))

    def exportMealPeriodViolations(self, stores, year_month_date, pay_period=0):
        text_csv = []
        mpvs = self.getMealPeriodViolations(stores, year_month_date, pay_period)
        grouped = defaultdict(list)
        for mpv in mpvs:
            grouped[f"{mpv['last_name']}, {mpv['first_name']}"].append(mpv)

        text_csv.append(
            "last_name,first_name,title,custom_earning_meal_period_violations,personal_note"
        )
        for g in grouped.values():
            mpv = g[0]
            t = (
                f"{mpv['last_name']},{mpv['first_name']},"
                + "Crew (Primary),"
                + f"{sum(item['hourly_rate'] for item in g)},"
                + f'"MPV {", ".join(list(str(item["day"].month) + "/" + str(item["day"].day) for item in g))}"'
            )
            if pay_period == 0:
                t += f',"{mpv["store"]}"'
            text_csv.append(t)
        return "\n".join(text_csv)

    def exportTipsTransform(self, tips_stream):
        text_csv = []
        workbook = openpyxl.load_workbook(tips_stream, data_only=True)
        text_csv.append("last_name,first_name,title,paycheck_tips")
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            if isinstance(worksheet, Worksheet):
                for row in islice(worksheet.rows, 3, None):
                    cell_value = row[3].value
                    if cell_value is not None:
                        try:
                            decimal_value = Decimal(str(cell_value)).quantize(
                                TWO_PLACES
                            )
                            if decimal_value > 0:
                                text_csv.append(
                                    f"{row[0].value},{row[1].value},Crew (Primary),{decimal_value}"
                                )
                        except (ValueError, InvalidOperation):
                            logger.exception(
                                "Could not convert value",
                                extra=dict(
                                    cell_value=cell_value,
                                    row=row,
                                    sheet_name=sheet_name,
                                    type=type(cell_value),
                                ),
                            )
        return "\n".join(text_csv)

    def getUser(self, user_id: int) -> dict[str, Any] | None:
        user_response: dict[str, Any] | None = self._a.get("/users/{}".format(user_id))
        if user_response is None:
            return None
        else:
            return user_response["user"]
